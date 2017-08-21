# !/usr/bin/env python
import os
import logging
import time
import datetime
import glob
from openpyxl import load_workbook
from collections import defaultdict

from banks.base_bank import BaseBank
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By

script_path = os.path.abspath(__file__)
script_folder = os.path.dirname(script_path)
screenshots_folder = os.path.join(script_folder, 'Screesnshots')
downloads_folder = os.path.join(os.environ['USERPROFILE'], 'Downloads')


class Poalim(BaseBank):
    def __init__(self, driver, opt_dict, bank_dict):
        super(Poalim, self).__init__(driver, opt_dict, bank_dict)

    def navigate_to_login(self):
        logging.info('Navigate to login started')
        self.driver.get(self.login_url)
        WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.CLASS_NAME, 'clsLoginBox')))
        assert self.title in self.driver.title
        logging.info('Navigate to login finished')

    def login(self):
        logging.info('Login started')
        frame = self.driver.find_element_by_tag_name('iframe')
        self.driver.switch_to.frame(frame)
        #self.driver.find_element_by_id('userID').send_keys(self.opt_dict['user_name'])
        self.driver.find_element_by_id('userID').send_keys(''.join(self.opt_dict['cred'][1:][::3]))
        self.driver.find_element_by_id('userPassword').send_keys(self.opt_dict['pass'])
        self.driver.find_element_by_id('inputSend').click()

        body_text = ''
        for i in range(10):  # waiting for excel export production
            body_text = self.driver.find_element_by_tag_name('body').text
            if 'שלום' in body_text:
                logging.info('Login succeeded')
                break
            logging.debug('Waiting for page load')
            time.sleep(3)

        if 'שלום' not in body_text:
            self.driver.save_screenshot(os.path.join(screenshots_folder, 'login_failed.png'))
            raise RuntimeError('Login failed')
        elif self.opt_dict['account_no'] not in body_text or self.opt_dict['branch_no'] not in body_text:
            self.driver.save_screenshot(os.path.join(screenshots_folder, 'login_wrong.png'))
            raise ValueError('Wrong branch/account - %s/%s' % (self.opt_dict['branch_no'], self.opt_dict['account_no']))

        logging.info('Login finished')

    def navigate_to(self, page):
        pages_dict = {
            'last_trxs': {'url': 'currentaccount', 'by_type': 'class_name', 'by_val': 'bp-period-select'},
            'credit_cards': {'url': 'plasticCards', 'by_type': 'css', 'by_val': 'button.dropdown-toggle'},
            'loans': {'url': 'creditAndMortgage', 'by_type': 'class_name', 'by_val': 'piroutHaHescbonot'}
        }

        logging.info('Navigate to %s page' % page)
        self.driver.get("%s/%s" % (self.main_url, pages_dict[page]['url']))
        if 'by_type' in pages_dict[page] and 'by_val' in pages_dict[page]:
            if pages_dict[page]['by_type'] == 'class_name':
                WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.CLASS_NAME, pages_dict[page]['by_val'])))  # list of credit cards
            elif pages_dict[page]['by_type'] == 'id':
                    WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.ID, pages_dict[page]['by_val'])))  # list of credit cards
            elif pages_dict[page]['by_type'] == 'css':
                    WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.CSS_SELECTOR, pages_dict[page]['by_val'])))  # list of credit cards

    def show_last_trxs(self):
        logging.info('Showing last trxs started')
        self.driver.find_element_by_class_name('bp-period-select').click()  # open period drop down list
        self.driver.find_elements_by_class_name('dropdown-divider')[4].click()  # select date range option

        if self.opt_dict['date_from']:
            logging.info('Entering start date')
            self.driver.find_element_by_name('general.startDate').clear()
            self.driver.find_element_by_name('general.startDate').send_keys(self.opt_dict['date_from'].replace('/', '.'))
            self.driver.find_element_by_name('general.endDate').clear()
            self.driver.find_element_by_name('general.endDate').send_keys(self.opt_dict['date_to'].replace('/', '.'))
            self.driver.find_element_by_css_selector('button.btn3.pull-left').click()

        logging.info('Showing last trxs finished')

    def export_data_click(self):
        logging.info('Clicking on export results')
        WebDriverWait(self.driver, 10).until(ec.element_to_be_clickable((By.CSS_SELECTOR, 'button.dropdown-toggle')))  # list of credit cards
        time.sleep(3)
        self.driver.find_element_by_css_selector('a.icon.icon-excel').click()

    def export_data(self):
        self.export_data_click()

        files = []
        for i in range(10):  # waiting for excel export production
            files = glob.glob(os.path.join(downloads_folder, '*.xlsx'))
            if len(files) == 1:
                break
            time.sleep(1)
            logging.info('Exporting results to excel - %d seconds' % i)

        if len(files) == 0:
            raise RuntimeError('Failed to export data to excel file')
        logging.info('Exporting last trxs finished')
        return files[0]

    def read_last_trxs_from_excel(self, file):
        logging.info('Reading data from excel file')
        self.results_dict['last_trxs'] = defaultdict(dict)
        col_headers = []
        workbook = load_workbook(file)
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)
        headers_row = 5

        for row_num, row_data in enumerate(worksheet.iter_rows()):
            if row_num > headers_row:
                for col_num, col_data in enumerate(row_data):
                    col_data = col_data.value
                    if isinstance(col_data, datetime.date):
                        col_data = '%s/%s/%s' %(col_data.day, col_data.month, col_data.year)
                    self.results_dict['last_trxs'][row_num - headers_row]['%s' % col_headers[col_num]] = col_data
            elif row_num == headers_row:
                col_headers = [col_name.value for col_name in row_data]
        os.remove(os.path.join(downloads_folder, file))

    def read_credit_cards_charges_from_excel(self, file):
        logging.info('Reading data from excel file')
        #self.results_dict['credit_cards_charges'] = defaultdict(dict)
        col_headers = []
        workbook = load_workbook(file)
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)
        headers_row = 5

        for row_num, row_data in enumerate(worksheet.iter_rows()):
            if row_num > headers_row:
                self.results_dict['credit_cards_charges'] = row_data[0].value
                break
        os.remove(os.path.join(downloads_folder, file))

    def read_loans_from_excel(self, file):
        logging.info('Reading data from excel file')
        #self.results_dict['loans'] = defaultdict(dict)
        workbook = load_workbook(file)
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)
        headers_row = 5

        for row_num, row_data in enumerate(worksheet.iter_rows()):
            if row_num > headers_row:
                self.results_dict['loans'] = row_data[0].value
                break
        os.remove(os.path.join(downloads_folder, file))
